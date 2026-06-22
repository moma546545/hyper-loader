
from core.utils import get_app_data_dir
import csv
import json
import logging
import os
import glob
import re
import sys
import threading
from datetime import datetime

try:
    import winreg
except ImportError:
    winreg = None

try:
    from PySide6.QtCore import QTimer, QTime
    from PySide6.QtWidgets import QFileDialog, QLineEdit, QSpinBox, QCheckBox, QComboBox, QSystemTrayIcon
except ImportError:
    from PyQt6.QtCore import QTimer, QTime
    from PyQt6.QtWidgets import QFileDialog, QLineEdit, QSpinBox, QCheckBox, QComboBox, QSystemTrayIcon

from core.audio_normalizer import normalize_folder
from core.anti_detection import anti_detection_engine
from core.bandwidth_scheduler import scheduler
from core.config import DEFAULT_SETTINGS, THEME_MODE_MAP, default_download_dir, estimate_file_size_bytes
from core.cookie_importer import auto_detect_and_export, encrypt_cookie_file_inplace
from core.database import (
    close_thread_connection,
    fetch_history,
    get_all_stats,
    increment_stat,
    insert_history,
    load_queue_items,
    load_session_settings,
    record_peak_speed,
    save_queue_items,
    save_session_settings,
)
from core.downloader import DownloadWorker
from core.duplicate_finder import build_duplicate_report
from core.proxy_manager import proxy_manager
from core.storage_watchdog import format_bytes, has_enough_space
from core.sustainability import sustainability
from core.i18n import i18n, _
from core.error_handler import ErrorHandler
from core.task_types import DownloadTask
from core.workers import AnalyzeWorker
from ui.themes import THEMES

logger = logging.getLogger("SnapDownloader")
MAX_IMPORT_FILE_BYTES = 10 * 1024 * 1024


class ImportController:
    def __init__(self, window):
        self.window = window

    def bulk_import(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.window,
            _("Open Links File"),
            "",
            _("Import Files (*.txt *.json *.csv *.xlsx);;Text Files (*.txt);;JSON Files (*.json);;CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)"),
        )
        if not file_path:
            return
        self.window._set_status("جاري استيراد الروابط...")
        worker = threading.Thread(
            target=self.bulk_import_worker,
            args=(file_path,),
            daemon=True,
            name="BulkImportWorker",
        )
        worker.start()

    def parse_bulk_import_links(self, file_path: str) -> list[str]:
        self._validate_import_file_size(file_path)
        ext = os.path.splitext(str(file_path or "").lower())[1]
        if ext == ".csv":
            return self._parse_csv_links(file_path)
        if ext == ".json":
            return [str(entry.get("url", "")).strip() for entry in self._parse_json_entries(file_path) if str(entry.get("url", "")).strip()]
        if ext == ".xlsx":
            return self._parse_xlsx_links(file_path)
        links = []
        seen = set()
        with open(file_path, "r", encoding="utf-8", errors="replace") as file:
            for raw_line in file:
                line = str(raw_line or "").strip()
                if not line or not line.startswith(("http://", "https://")):
                    continue
                if line in seen:
                    continue
                seen.add(line)
                links.append(line)
        return links

    def _validate_import_file_size(self, file_path: str) -> None:
        size = os.path.getsize(file_path)
        if size > MAX_IMPORT_FILE_BYTES:
            raise ValueError(
                f"Import file is too large ({size} bytes). "
                f"Maximum allowed size is {MAX_IMPORT_FILE_BYTES} bytes."
            )

    def _parse_csv_links(self, file_path: str) -> list[str]:
        links: list[str] = []
        seen = set()
        with open(file_path, "r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not isinstance(row, dict):
                    continue
                url = str(row.get("url", "") or row.get("link", "") or "").strip()
                if not url.startswith(("http://", "https://")):
                    continue
                if url in seen:
                    continue
                seen.add(url)
                links.append(url)
        return links

    def _parse_xlsx_links(self, file_path: str) -> list[str]:
        try:
            import openpyxl
        except Exception:
            logger.warning("openpyxl is not installed. Falling back to no links for xlsx import.")
            return []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header = [str(v or "").strip().lower() for v in next(rows, [])]
            try:
                url_idx = header.index("url")
            except ValueError:
                try:
                    url_idx = header.index("link")
                except ValueError:
                    return []
            links: list[str] = []
            seen = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                url = str(row[url_idx] or "").strip() if url_idx < len(row) else ""
                if not url.startswith(("http://", "https://")):
                    continue
                if url in seen:
                    continue
                seen.add(url)
                links.append(url)
            return links
        finally:
            wb.close()

    def bulk_import_worker(self, file_path: str):
        try:
            entries = self.parse_bulk_import_entries(file_path)
            links = [str(e.get("url", "")) for e in entries if str(e.get("url", "")).strip()]
            self.window.bulk_import_finished_ui.emit(
                {
                    "ok": True,
                    "file_path": file_path,
                    "links": links,
                    "entries": entries,
                }
            )
        except (PermissionError, FileNotFoundError) as exc:
            logger.warning(f"Bulk import file error: {exc}")
            self.window.bulk_import_finished_ui.emit(
                {
                    "ok": False,
                    "file_path": file_path,
                    "error": f"Cannot read file: {exc}",
                    "links": [],
                    "entries": [],
                }
            )
        except (UnicodeDecodeError, ValueError) as exc:
            logger.exception("Bulk import failed to decode or parse file")
            self.window.bulk_import_finished_ui.emit(
                {
                    "ok": False,
                    "file_path": file_path,
                    "error": str(exc),
                    "links": [],
                    "entries": [],
                }
            )

    def parse_bulk_import_entries(self, file_path: str) -> list[dict]:
        self._validate_import_file_size(file_path)
        ext = os.path.splitext(str(file_path or "").lower())[1]
        if ext == ".csv":
            return self._parse_csv_entries(file_path)
        if ext == ".json":
            return self._parse_json_entries(file_path)
        if ext == ".xlsx":
            return self._parse_xlsx_entries(file_path)
        return [{"url": u} for u in self.parse_bulk_import_links(file_path)]

    def handle_bulk_import_result(self, payload: dict):
        self.window._set_status("جاهز")
        if not payload.get("ok"):
            self.window._record_error(
                str(payload.get("error", "Unknown error") or "Unknown error"),
                title="Bulk Import",
                source="bulk_import",
            )
            ErrorHandler.show_error(
                self.window,
                _("Error"),
                _("Failed to import links: {error}").format(
                    error=payload.get("error", "Unknown error")
                ),
            )
            return
        links = list(payload.get("links") or [])
        entries = list(payload.get("entries") or [])
        if not links:
            ErrorHandler.show_warning(
                self.window,
                _("No Links"),
                _("No valid links found in the selected file."),
            )
            return
        logger.info(f"Imported {len(links)} links from {payload.get('file_path', '')}")
        tasks: list[DownloadTask] = []
        raw_entries = entries if entries else [{"url": u} for u in links]
        for entry in raw_entries:
            if not isinstance(entry, dict):
                continue
            link = str(entry.get("url", "")).strip()
            if not link:
                continue
            task = self.window._normalize_task(
                self.window._build_task(url=link, title=f"Bulk: {link}"),
                title=f"Bulk: {link}",
            )
            if entry.get("format"):
                task["format"] = str(entry.get("format", task.get("format", ""))).strip()
            if entry.get("quality"):
                task["quality"] = str(entry.get("quality", task.get("quality", ""))).strip()
            if entry.get("subtitle"):
                task["subtitle"] = str(entry.get("subtitle", task.get("subtitle", "None"))).strip()
            if entry.get("out_dir"):
                task["out_dir"] = str(entry.get("out_dir", task.get("out_dir", ""))).strip() or task.get("out_dir", "")
            if entry.get("post_action"):
                task["post_action"] = str(entry.get("post_action", "none")).strip() or "none"
            if entry.get("post_download_script"):
                task["post_download_script"] = str(entry.get("post_download_script", "")).strip()
            if entry.get("post_process_pipeline"):
                task["post_process_pipeline"] = list(entry.get("post_process_pipeline", []) or [])
            if entry.get("video_id"):
                task["video_id"] = str(entry.get("video_id", "")).strip()
            if entry.get("bandwidth_limit_kbps") not in {None, ""}:
                try:
                    task["bandwidth_limit_kbps"] = max(0, int(entry.get("bandwidth_limit_kbps")))
                except Exception:
                    pass
            tasks.append(task)
        self.window.queue_manager.add_tasks(tasks)
        ErrorHandler.show_info(
            self.window,
            _("Import Success"),
            _("Successfully imported {count} links to the queue.").format(count=len(links)),
        )
        self.window._switch_view("downloads")
        self.window._start_queue_download()

    def _parse_csv_entries(self, file_path: str) -> list[dict]:
        entries: list[dict] = []
        with open(file_path, "r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not isinstance(row, dict):
                    continue
                url = str(row.get("url", "") or row.get("link", "") or "").strip()
                if not url.startswith(("http://", "https://")):
                    continue
                entries.append(
                    {
                        "url": url,
                        "title": str(row.get("title", "") or "").strip(),
                        "format": str(row.get("format", "") or "").strip(),
                        "quality": str(row.get("quality", "") or "").strip(),
                        "subtitle": str(row.get("subtitle", "") or "").strip(),
                        "out_dir": str(row.get("out_dir", "") or "").strip(),
                        "post_action": str(row.get("post_action", "") or "").strip(),
                        "post_download_script": str(row.get("post_download_script", "") or "").strip(),
                        "video_id": str(row.get("video_id", "") or "").strip(),
                        "bandwidth_limit_kbps": str(row.get("bandwidth_limit_kbps", "") or "").strip(),
                    }
                )
        return entries

    def _parse_xlsx_entries(self, file_path: str) -> list[dict]:
        try:
            import openpyxl
        except Exception:
            logger.warning("openpyxl is not installed. Falling back to plain links for xlsx import.")
            return []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            header = [str(v or "").strip().lower() for v in header_cells]
            idx = {name: i for i, name in enumerate(header)}
            url_idx = idx.get("url", idx.get("link", -1))
            if url_idx < 0:
                return []
            entries: list[dict] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                url = str(row[url_idx] or "").strip() if url_idx < len(row) else ""
                if not url.startswith(("http://", "https://")):
                    continue

                def _v(col: str) -> str:
                    i = idx.get(col, -1)
                    if i < 0 or i >= len(row):
                        return ""
                    return str(row[i] or "").strip()

                entries.append(
                    {
                        "url": url,
                        "title": _v("title"),
                        "format": _v("format"),
                        "quality": _v("quality"),
                        "subtitle": _v("subtitle"),
                        "out_dir": _v("out_dir"),
                        "post_action": _v("post_action"),
                        "post_download_script": _v("post_download_script"),
                        "video_id": _v("video_id"),
                        "bandwidth_limit_kbps": _v("bandwidth_limit_kbps"),
                    }
                )
            return entries
        finally:
            wb.close()

    def _parse_json_entries(self, file_path: str) -> list[dict]:
        with open(file_path, "r", encoding="utf-8", errors="replace") as file:
            payload = json.load(file)
        raw_entries = payload if isinstance(payload, list) else payload.get("entries", payload.get("links", []))
        entries: list[dict] = []
        for raw in raw_entries or []:
            if isinstance(raw, str):
                url = raw.strip()
                if url.startswith(("http://", "https://")):
                    entries.append({"url": url})
                continue
            if not isinstance(raw, dict):
                continue
            url = str(raw.get("url", "") or raw.get("link", "") or "").strip()
            if not url.startswith(("http://", "https://")):
                continue
            entries.append(
                {
                    "url": url,
                    "title": str(raw.get("title", "") or "").strip(),
                    "format": str(raw.get("format", "") or "").strip(),
                    "quality": str(raw.get("quality", "") or "").strip(),
                    "subtitle": str(raw.get("subtitle", "") or "").strip(),
                    "out_dir": str(raw.get("out_dir", "") or "").strip(),
                    "post_action": str(raw.get("post_action", "") or "").strip(),
                    "post_download_script": str(raw.get("post_download_script", "") or "").strip(),
                    "video_id": str(raw.get("video_id", "") or "").strip(),
                    "bandwidth_limit_kbps": str(raw.get("bandwidth_limit_kbps", "") or "").strip(),
                    "post_process_pipeline": list(raw.get("post_process_pipeline", []) or []),
                }
            )
        return entries


